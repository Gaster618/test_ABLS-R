from openpyxl import load_workbook

name = input('Введите фамилию: ')
wb = load_workbook('123.xlsx')
sheet =  wb['Sheet1']
for i in range(1, sheet.max_row + 1):
	if sheet.cell(row=i, column=2).value == name:
		for cell in range(1, sheet.max_column + 1):
			print(sheet.cell(row=i, column=cell).value)