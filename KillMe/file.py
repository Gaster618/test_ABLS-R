import flask
from openpyxl import load_workbook

app = flask.Flask(__name__)

@app.route('/', strict_slashes=False, methods=['GET', 'POST'])
def index():
	if flask.request.method == 'GET':
		return flask.render_template('index.html')
	else:
		wb = load_workbook('smth.xlsx')
		sheet = wb['Sheet1']
		sheet.cell(row=sheet.max_row + 1, column=1).value = flask.request.form["tea"]
		sheet.cell(row=sheet.max_row, column=2).value = flask.request.form["glass"]
		wb.save('smth.xlsx')
		return 'Ваши ответы сохранены'

app.run()